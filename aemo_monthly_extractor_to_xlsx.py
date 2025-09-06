import requests
import datetime
import os
import shutil
import csv
import time
from openpyxl import Workbook

CONST_URL: str = 'https://aemo.com.au/aemo/data/nem/priceanddemand/'

# AEMO thinks it's smart, but it's not if they require this
url_headers = {
  'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/58.0.3029.110 Safari/537.36'
}

# It was 2 cases of this code, so I made function for this.
def recreate(dir_name: str) -> None:
  if os.path.exists(dir_name):
    shutil.rmtree(dir_name)
  os.makedirs(dir_name)

# The Application
def app():
  recreate('data')
  ct_data: str = datetime.date.today().strftime("%Y%m")
  url: str = ''
  csv_files: list = []
  for nem in ['NSW1', 'QLD1', 'VIC1', 'SA1', 'TAS1']:
    file_name = f'PRICE_AND_DEMAND_{ct_data}_{nem}.csv'
    csv_files.append(file_name)
    print(f'Download: {CONST_URL}{file_name}')
    req = requests.get(f'{CONST_URL}{file_name}', headers=url_headers)
    with open(f'data/{file_name}', 'wb') as file:
      file.write(req.content)
      
  workbook = Workbook()
  ct_sheet = workbook.active    

  ct_column, ct_raw = 1, 1
  title, value = str(), float() #Initialize here for loop optimization
  for csv_file in csv_files:
    with open(f'data/{csv_file}', 'r') as csvfile:
        csv_reader = csv.reader(csvfile)
        next(csv_reader)
        for row in csv_reader:
          if ct_raw == 1:
            title = row[0] # For Logging
            ct_sheet.cell(row=ct_raw, column=ct_column, value=title)
          ct_raw += 1
          value = float(row[3]) # For Logging
          ct_sheet.cell(row=ct_raw, column=ct_column, value=value)
          print(f'NEM: {title}\tRAW: {ct_raw}\tCOLUMN: {ct_column}\tVALUE: {value}') #Log
    ct_raw = 1
    ct_column += 1
    title, value = str(), float()
    print()
  
  print('Time: ', datetime.datetime.now())
  recreate('result')
  workbook.save('result/result.xlsx')
        
if __name__ == '__main__':
  try:
    app()
    while True:
      today = datetime.datetime.now()
      # Fancy way to check 00:00 on first day of a month
      if [today.day, today.hour, today.minute] == [1, 0, 0]:
        app()
      else:
        time.sleep(1)
  except KeyboardInterrupt:
    print('[Ctrl+C] Exit from a program')
  except Exception as e:
    print('Unknown Error')